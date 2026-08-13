#!/usr/bin/env python3
"""Assemble the BS2 review workbook from the gene table, the misfire detail
TSV and the re-annotated misfire JSON."""
import csv
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

GENE_TSV, DETAIL_TSV, MISFIRE_JSON, OUT = sys.argv[1:5]

HEAD = PatternFill("solid", fgColor="1F3864")
REVIEW_HEAD = PatternFill("solid", fgColor="7B3F00")
BAND = PatternFill("solid", fgColor="F2F2F2")
WHITE = Font(color="FFFFFF", bold=True, size=10)

REVIEW_PREFIXES = ("BS2_applicable", "BS1_applicable", "onset__", "penetrance__",
                   "variable_expressivity", "disease_prevalence",
                   "recommended_", "founder_population",
                   "hypomorph_in_trans", "reviewer_notes")


def is_review(col):
    return col.startswith(REVIEW_PREFIXES)


def style(ws, header, widths):
    for i, col in enumerate(header, 1):
        c = ws.cell(row=1, column=i)
        c.fill = REVIEW_HEAD if is_review(col) else HEAD
        c.font = WHITE
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "B2"
    for r in range(2, ws.max_row + 1):
        if r % 2 == 0:
            for i in range(1, len(header) + 1):
                ws.cell(row=r, column=i).fill = BAND
        for i in range(1, len(header) + 1):
            ws.cell(row=r, column=i).alignment = Alignment(vertical="top", wrap_text=True)


wb = Workbook()

# ── Sheet 1: gene-level ─────────────────────────────────────────────────────
rows = list(csv.DictReader(open(GENE_TSV), delimiter="\t"))
# Everything with a misfire, plus the highest-volume genes, capped so the
# reviewer sees a list comparable in size to the 122 she worked through.
misfire_rows = [r for r in rows if int(r["n_on_pathogenic_truth"]) > 0]
rest = [r for r in rows if int(r["n_on_pathogenic_truth"]) == 0]
rest.sort(key=lambda r: -int(r["n_variants_with_BS2"]))
selected = misfire_rows + rest[:70]

ws = wb.active
ws.title = "BS2 by gene"
header = list(rows[0].keys())
ws.append(header)
for r in selected:
    ws.append([r.get(c, "") for c in header])
style(ws, header, {
    "gene": 12, "clinvar_diseases_top3": 46, "example_misfires": 52,
    "most_common_criteria_signature": 26, "reviewer_notes": 40,
    "n_variants_with_BS2": 11, "n_on_pathogenic_truth": 11,
    "n_on_benign_truth": 11, "n_on_uncertain_truth": 11,
})

# ── Sheet 2: variant-level misfires ─────────────────────────────────────────
# Matched by record order, not by coordinate: the annotator preserves input
# order, and a coordinate key fails on indels because the VCF carries the
# anchor base (`CA>C`) while the annotator reports the trimmed allele (`A/-`).
detail_rows = list(csv.DictReader(open(DETAIL_TSV), delimiter="\t"))

doc = json.load(open(MISFIRE_JSON))
recs = doc if isinstance(doc, list) else doc.get("variants", [])
assert len(recs) == len(detail_rows), (
    f"{len(recs)} annotated records vs {len(detail_rows)} detail rows; "
    "order-based matching is only valid when they correspond one to one")
ws2 = wb.create_sheet("BS2 misfires (variants)")
h2 = ["gene", "clinvar_hgvs_g", "consequence", "clinvar_call", "fastvep_call",
      "criteria", "clinvar_disease",
      "gnomad_homozygotes", "gnomad_individuals", "gnomad_AC",
      "hom_freq_95pct_lower_bound", "prevalence_bar", "inheritance_we_inferred",
      "why_BS2_fired"] + [c for c in rows[0] if is_review(c)]
ws2.append(h2)

for r, d in zip(recs, detail_rows):
    for tc in (r.get("transcript_consequences") or [])[:1]:
        bs2 = next((c for c in (tc.get("acmg") or {}).get("criteria", [])
                    if c["code"] == "BS2"), None)
        det = (bs2 or {}).get("details") or {}
        inh = ("recessive" if det.get("omim_recessive") else
               "dominant" if det.get("omim_dominant") else "not stated in GDV")
        lb = det.get("hom_freq_lower_95")
        ws2.append([
            d.get("gene") or tc.get("gene_symbol", ""),
            d.get("hgvs", ""), d.get("consequence", ""),
            d.get("clnsig", ""), (tc.get("acmg") or {}).get("shorthand", ""),
            d.get("criteria", ""), d.get("clndn", ""),
            det.get("gnomad_allHc", ""), det.get("gnomad_individuals", ""),
            det.get("gnomad_allAc", ""),
            f"{lb:.2e}" if isinstance(lb, float) else "",
            det.get("prevalence_threshold", ""), inh,
            (bs2 or {}).get("summary", ""),
        ] + [""] * sum(1 for c in rows[0] if is_review(c)))
style(ws2, h2, {
    "gene": 12, "clinvar_hgvs_g": 30, "consequence": 24,
    "clinvar_call": 18, "fastvep_call": 12, "criteria": 26,
    "clinvar_disease": 40, "why_BS2_fired": 62, "reviewer_notes": 40,
    "inheritance_we_inferred": 16,
})

# ── Sheet 3: key ────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("KEY")
ws3.append(["What BS2 uses today, and what this table is for"])
ws3["A1"].font = Font(bold=True, size=13)
lines = [
    "",
    "There is no curated disease list behind BS2. It decides per variant from four inputs:",
    "  1. gnomAD counts (homozygotes; hemizygotes on non-PAR chrX; allele count).",
    "  2. Inheritance, parsed as free text out of the ClinGen Gene-Disease Validity table",
    "     ('autosomal recessive' / 'autosomal dominant'). Where GDV does not state it, BS2",
    "     falls back to the recessive test. That fallback is worth your attention.",
    "  3. One hard-coded gene list, of 25 genes with a pseudogene or near-identical paralogue",
    "     (Mandelker 2016). It BLOCKS BS2 entirely; it does not tune it.",
    "     CYP21A2, STRC, HBA1, HBA2, SMN1, SMN2, PMS2, NEB, OTOA, GBA, IKBKG, CFC1, NCF1,",
    "     TTN, CYP11B1, CYP11B2, HBB, HBD, SBDS, FCGR3A, FCGR3B, CR1, C4A, C4B, TUBB8,",
    "     MOCS1, OPN1LW, OPN1MW, GTF2I",
    "  4. ClinVar's own 'low penetrance' / 'risk allele' terms, which also block BS2.",
    "",
    "Two numeric thresholds, applied to every gene on earth identically:",
    "  - Recessive / X-linked: at least 2 individuals with no functional copy, AND the 95%",
    "    lower bound on their frequency must exceed 1 in 1,000. That bar is a maximum credible",
    "    disease prevalence, and we set it by sweeping the whole benchmark rather than by",
    "    convention: it has to cover the most prevalent Mendelian conditions, which is why it",
    "    sits at the prevalence of hearing loss and alpha-1 antitrypsin deficiency rather than",
    "    at that of a typical rare disease.",
    "  - Dominant: allele count of at least 5 in unaffected adults.",
    "",
    "WHAT WE ARE ASKING YOU FOR",
    "",
    "The flat 1-in-100,000 bar is the weakest part of this. Your email said BS1 and BS2 may not",
    "apply at all for common Mendelian conditions (hearing loss, alpha-1 antitrypsin deficiency,",
    "familial Mediterranean fever), for late-onset disease, and where penetrance is low or",
    "expressivity variable. We cannot encode that without a per-gene-disease statement from",
    "someone who knows the diseases. These sheets are the seed for that table.",
    "",
    "The brown columns are yours. Please fill in what you can and leave the rest blank:",
    "  BS2_applicable / BS1_applicable  - can these criteria be used for this disease at all?",
    "  onset                            - congenital / childhood / adult / late",
    "  penetrance                       - complete / incomplete / unknown",
    "  variable_expressivity            - Y / N",
    "  disease_prevalence               - your best figure, any format ('1 in 50,000', '~2%')",
    "  recommended_BA1 / BS1 threshold  - if you know the VCEP value, or would set one",
    "  founder_population               - if the allele is enriched in a specific population",
    "  hypomorph_in_trans_mechanism     - the TAR-type mechanism from your email: a common",
    "                                     hypomorph that only causes disease in trans with a",
    "                                     LoF allele, and is harmless when homozygous",
    "  reviewer_notes                   - anything else",
    "",
    "SHEETS",
    "",
    "  'BS2 by gene'             - one row per gene. Sorted so that genes where BS2 fired on a",
    "                              variant ClinVar calls pathogenic come first: those are where",
    "                              the criterion is actively doing harm. Below them, the",
    "                              highest-volume genes, for context.",
    "  'BS2 misfires (variants)' - the 109 individual variants where BS2 fired against a",
    "                              pathogenic ClinVar call, with the actual gnomAD homozygote",
    "                              count, the cohort size, and the computed lower bound, so you",
    "                              can see exactly what the criterion was reasoning from.",
    "",
    "Counts are from benchmark run v9 over all 673,660 ClinVar 2-star-or-better variants.",
    "BS2 fired 67,615 times in total; it fired on pathogenic truth 109 times, in 31 genes.",
    "",
    "WHERE THE HARM ACTUALLY IS",
    "",
    "Of those 109, 79 came from the DOMINANT branch - allele count of 5 or more in unaffected",
    "adults - and only 30 from the recessive/hemizygous test. We had assumed the homozygote",
    "rule was the main problem; it is not. A flat allele count of 5 in a cohort of 730,000 is",
    "very little evidence for a late-onset dominant condition, which is why APOB (familial",
    "hypercholesterolemia, 17 rows), PKD2 (adult polycystic kidney disease, 8) and TERT (4)",
    "sit near the top. BEST1 leads with 39, where incomplete penetrance is the issue.",
    "",
    "So the onset and penetrance columns matter as much for dominant conditions as for",
    "recessive ones. If a dominant disease presents in adulthood, please say so even where",
    "the homozygote question does not arise.",
    "",
    "One more thing we would value your eye on: for 21 of the 109 our gene-disease table",
    "states no inheritance at all, and BS2 silently falls back to the recessive test",
    "(PADI3 and MUTYH are examples). Those rows are marked 'not stated in GDV'.",
]
for ln in lines:
    ws3.append([ln])
ws3.column_dimensions["A"].width = 104
for r in range(1, ws3.max_row + 1):
    ws3.cell(row=r, column=1).alignment = Alignment(vertical="top")

wb.save(OUT)
print(f"wrote {OUT}: {len(selected)} genes, {ws2.max_row - 1} misfire variants")
