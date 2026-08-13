#!/usr/bin/env python3
"""Append fastVEP's current call to a reviewer's own annotated workbook.

The medical geneticist annotates the discordance table we send and returns it.
Sending back a freshly generated table would lose that work and force her to
re-locate every row she has already reasoned about. This instead opens her file,
re-annotates exactly the variants in it, and appends new columns on the right,
leaving her notes, her colour coding and her row order untouched.

Appended columns:

  fastvep_now_class       the current call
  fastvep_now_criteria    criteria met, `&`-joined
  changed_since_review    what happened to the discordance she reviewed
  what_changed            which rule moved it, in one phrase

Usage:
  05_append_current_calls.py <reviewed.xlsx> <out.xlsx> [--sheet Sheet1]
                            [--acmg-config <file.toml>]
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
ROOT = os.path.dirname(ROOT)

NEW_COLUMNS = [
    "fastvep_now_class",
    "fastvep_now_criteria",
    "changed_since_review",
    "what_changed",
]

BENIGN = {"B", "LB"}
PATHOGENIC = {"P", "LP"}

LONG_NAME = {
    "P": "Pathogenic",
    "LP": "Likely_pathogenic",
    "VUS": "VUS",
    "LB": "Likely_benign",
    "B": "Benign",
}


def bucket(label: str) -> str:
    """Collapse a class label (long or short form) to a direction."""
    if not label:
        return "uncertain"
    s = str(label).strip()
    if s in PATHOGENIC or s in ("Pathogenic", "Likely_pathogenic"):
        return "pathogenic"
    if s in BENIGN or s in ("Benign", "Likely_benign"):
        return "benign"
    return "uncertain"


def read_rows(ws, header):
    idx = {name: i for i, name in enumerate(header) if name}
    need = ("chrom", "pos", "ref", "alt")
    missing = [c for c in need if c not in idx]
    if missing:
        sys.exit(f"reviewed workbook is missing column(s): {', '.join(missing)}")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[idx["chrom"]] is None:
            rows.append(None)  # keep the blank row so output stays aligned
            continue
        rows.append(
            {
                "chrom": str(r[idx["chrom"]]).strip(),
                "pos": str(r[idx["pos"]]).strip(),
                "ref": str(r[idx["ref"]]).strip(),
                "alt": str(r[idx["alt"]]).strip(),
                "truth": str(r[idx.get("truth_class", -1)] or "") if "truth_class" in idx else "",
                "then": str(r[idx.get("fastvep_class", -1)] or "") if "fastvep_class" in idx else "",
                "then_criteria": (
                    str(r[idx["fastvep_met_criteria"]] or "")
                    if "fastvep_met_criteria" in idx
                    else ""
                ),
            }
        )
    return rows


def annotate(rows, workdir, acmg_config=None):
    """Run the current binary over exactly these variants, in this order."""
    vcf = os.path.join(workdir, "reviewed.vcf")
    with open(vcf, "w") as f:
        f.write("##fileformat=VCFv4.2\n#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\n")
        for r in rows:
            if r is None:
                continue
            # `ALT=.` is a ClinVar reference-agreement record, not a variant.
            # fastVEP skips those, so emitting one would silently shift every
            # subsequent record when we match results back by order.
            alt = r["alt"] if r["alt"] not in (".", "") else "N"
            f.write(f"{r['chrom']}\t{r['pos']}\t.\t{r['ref']}\t{alt}\t.\t.\t.\n")

    out = os.path.join(workdir, "reviewed.json")
    subprocess.run(
        [
            os.path.join(ROOT, "target/release/fastvep"), "annotate",
            "-i", vcf, "-o", out,
            "--gff3", os.path.join(ROOT, "test_data/organisms/human/Homo_sapiens.GRCh38.115.gff3"),
            "--fasta", os.path.join(ROOT, "test_data/organisms/human/Homo_sapiens.GRCh38.dna.primary_assembly.fa"),
            "--sa-dir", os.path.join(ROOT, "data/benchmark/sa_db"),
            "--acmg", "--pick", "--hgvs", "--output-format", "json",
        ]
        + (["--acmg-config", acmg_config] if acmg_config else []),
        check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    doc = json.load(open(out))
    recs = doc if isinstance(doc, list) else doc.get("variants", [])
    live = [r for r in rows if r is not None]
    if len(recs) != len(live):
        sys.exit(
            f"annotator returned {len(recs)} records for {len(live)} input variants; "
            "order-based matching is only valid one-to-one"
        )
    return recs


def summarise(rec):
    """(class, criteria, per-criterion state) for one annotated record."""
    tcs = rec.get("transcript_consequences") or []
    if not tcs:
        return "", "", []
    acmg = tcs[0].get("acmg") or {}
    all_criteria = acmg.get("criteria") or []
    met = [c for c in all_criteria if c.get("met")]
    return (
        acmg.get("shorthand", ""),
        "&".join(c["code"] for c in met),
        [
            (c["code"], c.get("summary", ""), bool(c.get("met")), bool(c.get("evaluated")))
            for c in all_criteria
        ],
    )


def explain(then_class, then_criteria, now_class, now_criteria, all_criteria):
    """How the row moved, and why, from the evidence rather than from a guess.

    `what_changed` is a diff of the criteria that actually fired, plus the named
    gate for any criterion we can see was explicitly withheld. Earlier this
    scanned every criterion's summary for gate phrases, which reported
    "filtering allele frequency" on rows where BA1 and BS1 had not fired at all
    - true of the code path, useless to a reader.
    """
    before = set(c for c in (then_criteria or "").replace(";", "&").split("&") if c)
    after = set(c for c in (now_criteria or "").split("&") if c)

    status = "unchanged"
    if then_class and now_class:
        b_then, b_now = bucket(then_class), bucket(now_class)
        if b_then != b_now:
            status = "direction changed"
        elif LONG_NAME.get(now_class, now_class) != str(then_class):
            status = "call changed"

    parts = []
    gained, lost = sorted(after - before), sorted(before - after)
    if gained:
        parts.append("gained " + ", ".join(gained))
    if lost:
        parts.append("lost " + ", ".join(lost))

    # Name the gate only for criteria that were explicitly withheld, i.e. the
    # evaluator looked and declined. `evaluated: false` with a gate phrase is
    # the signature; a criterion that simply did not meet its threshold is not
    # interesting here.
    gates = []
    for code, summary, met, evaluated in all_criteria:
        if met or evaluated:
            continue
        for needle, phrase in (
            ("no_valid_gene_disease_relationship", "gene-disease validity gate"),
            ("mechanism_not_loss_of_function", "gain-of-function mechanism gate"),
            ("gnomAD FILTER=", "gnomAD QC gate"),
            ("segmental duplication", "segdup region gate"),
            ("low-complexity region", "low-complexity region gate"),
            ("paralogue/pseudogene homology", "homologous-gene list"),
            ("low-penetrance", "ClinVar low-penetrance term"),
        ):
            if needle in (summary or ""):
                gates.append(f"{code} withheld: {phrase}")
                break
    for code, summary, met, _ in all_criteria:
        if not met and "Superseded by functional evidence" in (summary or ""):
            gates.append(f"{code} superseded by functional evidence")
    # The curated exception list reports `evaluated: true` - it is a
    # conclusion, not an inability to conclude - so it needs its own branch
    # rather than the withheld-criteria loop above. It now covers BS1 and BS2
    # as well as BA1, for hypomorphic alleles.
    for code, summary, met, _ in all_criteria:
        if not met and "frequency-exception list" in (summary or ""):
            gates.append(f"{code} blocked: curated frequency-exception list")
    parts.extend(dict.fromkeys(gates))

    if "PVS1" in after and not (after - {"PVS1"}) and status != "unchanged":
        parts.append("lone PVS1 reaches LP under the SVI point system")

    return status, "; ".join(parts)[:300]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("reviewed")
    ap.add_argument("out")
    ap.add_argument("--sheet", default=None, help="defaults to the first sheet")
    ap.add_argument(
        "--acmg-config",
        default=None,
        help="optional --acmg-config TOML, to show what a candidate configuration "
             "would do to the reviewed rows (e.g. the published VCEP frequency bars)",
    )
    args = ap.parse_args()

    wb = load_workbook(args.reviewed)
    sheet = args.sheet or wb.sheetnames[0]
    ws = wb[sheet]
    header = [c.value for c in ws[1]]

    rows = read_rows(ws, header)
    with tempfile.TemporaryDirectory() as workdir:
        recs = annotate(rows, workdir, args.acmg_config)

    # Style the appended headers so it is obvious which columns are new.
    fill = PatternFill("solid", fgColor="1F6F43")
    font = Font(color="FFFFFF", bold=True, size=10)
    first_new = ws.max_column + 1
    for offset, name in enumerate(NEW_COLUMNS):
        c = ws.cell(row=1, column=first_new + offset, value=name)
        c.fill, c.font = fill, font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = 46 if name == "what_changed" else 22

    it = iter(recs)
    changed = same = 0
    for i, row in enumerate(rows, start=2):
        if row is None:
            continue
        now, criteria, all_criteria = summarise(next(it))
        status, why = explain(
            row["then"], row["then_criteria"], now, criteria, all_criteria
        )
        if status == "unchanged":
            same += 1
        else:
            changed += 1
        for offset, value in enumerate([now, criteria, status, why]):
            cell = ws.cell(row=i, column=first_new + offset, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(args.out)
    print(f"wrote {args.out}")
    print(f"  {changed} row(s) changed since the reviewed run, {same} unchanged")


if __name__ == "__main__":
    main()
