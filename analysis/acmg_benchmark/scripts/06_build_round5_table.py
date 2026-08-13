#!/usr/bin/env python3
"""Build the round-5 medical-genetics review workbook.

One file, one sheet of variants: every call where fastVEP still commits to a
direction that contradicts ClinVar 2-star+. Rows the reviewer has already
resolved are gone by construction, because they are no longer discordant.

Every annotation she has made across the review is carried onto the row it
belongs to, one column per round, so nothing has to be re-located and nothing
is silently merged:

  md_round2_verdict   her ruling on the 122-row table (colour-coded, decoded)
  md_round2_note      her free text on that table
  md_round3_note      the 46-row table
  md_round4_note      the 122-row table as returned most recently

Rounds 3 and 4 carried her round-2 text forward rather than adding to it, so
those columns are near-identical in practice. They are emitted separately
anyway: collapsing them would be a judgement about her notes rather than a
presentation of them.

Usage:
  06_build_round5_table.py <review.tsv> <out.xlsx>
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[3]

# Where each round's annotations live. Order is the order of the columns.
ROUND2_TSV = ROOT / "data/benchmark/output_v8/md_review_round2_annotated.tsv"
ROUND3_XLSX = ROOT / "data/benchmark/output_v9/discordant_46_round3.xlsx"
ROUND4_XLSX = ROOT / "data/benchmark/discordance_review/discordant_122_round4.xlsx"

# Round-2 legend. Her ruling is recorded as a cell fill in the workbook she
# returns and as a text column in the round-2 TSV; the two agree, and the TSV
# is preferred because it needs no decoding.
VERDICT_LABEL = {
    "clinvar_correct": "ClinVar correct; fastVEP wrong",
    "fastvep_correct": "fastVEP correct; ClinVar wrong",
    "both_wrong": "both calls discordant with the reviewer's call",
    "needs_literature": "needs literature evidence",
}

# Columns the reviewer fills in this round. The last two are new and exist
# because her round-5 comments are about exactly those two things: whether a
# frequency criterion may be applied to a given disorder at all, and whether
# published functional data should override a prediction.
REVIEW_COLUMNS = [
    ("MD_CALL", "P / LP / VUS / LB / B - your classification for this variant"),
    ("WHO_IS_RIGHT", "clinvar / fastvep / neither"),
    ("REASON_FOR_DISCORDANCE", "free text; what evidence decides it"),
    ("ACMG_CODE_MISAPPLIED", "e.g. BS1, PVS1, BP7 - which code fastVEP got wrong"),
    (
        "BS1_BS2_NOT_APPLICABLE_WHY",
        "if frequency evidence should not apply here: common phenotype / late onset / "
        "reduced penetrance / variable expressivity / founder allele / hypomorph in trans",
    ),
    ("FUNCTIONAL_OR_LITERATURE_PMID", "PMID for functional or segregation data that settles it"),
]

CARRIED_COLUMNS = [
    "rounds_seen",
    "md_round2_verdict",
    "md_round2_note",
    "md_round3_note",
    "md_round4_note",
]

# Identifying columns, pulled to the front so a row can be located at a glance.
LEAD_COLUMNS = [
    "gene", "variant", "fastvep_HGVSc", "fastvep_HGVSp", "stars",
    "truth_class", "fastvep_class", "fastvep_met_criteria",
]

HEADER_FILL = PatternFill("solid", start_color="FF1F3864")
REVIEW_FILL = PatternFill("solid", start_color="FFFFF2CC")
CARRIED_FILL = PatternFill("solid", start_color="FFEAF1F8")


def key_of(values: dict) -> tuple[str, ...]:
    return tuple(str(values[c]).strip() for c in ("chrom", "pos", "ref", "alt"))


def load_round2() -> dict[tuple[str, ...], tuple[str, str]]:
    """(verdict, note) per variant, from the round-2 annotated TSV."""
    if not ROUND2_TSV.exists():
        return {}
    out = {}
    with ROUND2_TSV.open() as f:
        for r in csv.DictReader(f, delimiter="\t"):
            verdict = VERDICT_LABEL.get((r.get("reviewer_verdict") or "").strip(), "")
            note = (r.get("REASON FOR DISCORDANCE") or "").strip()
            out[key_of(r)] = (verdict, note)
    return out


def load_notes(path: Path, sheet: str, column: str) -> dict[tuple[str, ...], str]:
    """Free-text notes per variant from one of the returned workbooks."""
    if not path.exists():
        return {}
    ws = load_workbook(path)[sheet]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header) if name}
    if column not in idx:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["chrom"]] is None:
            continue
        values = {c: row[idx[c]] for c in ("chrom", "pos", "ref", "alt")}
        out[key_of(values)] = (row[idx[column]] or "").strip()
    return out


def write_workbook(path: Path, rows: list[dict[str, str]], header: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(header)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    review_names = {name for name, _ in REVIEW_COLUMNS}
    for row in rows:
        ws.append([row.get(name, "") for name in header])

    for i, name in enumerate(header, start=1):
        letter = get_column_letter(i)
        if name in review_names:
            width = 26
        elif name.startswith("md_round"):
            width = 34
        elif name in ("gene", "stars", "rounds_seen"):
            width = 11
        elif name in ("truth_class", "fastvep_class"):
            width = 17
        else:
            width = 22
        ws.column_dimensions[letter].width = width
        fill = REVIEW_FILL if name in review_names else (
            CARRIED_FILL if name in CARRIED_COLUMNS else None)
        if fill:
            for cell in ws[letter][1:]:
                cell.fill = fill

    # Dropdowns on the two columns with a closed vocabulary. Everything else is
    # free text on purpose: the useful part of her review is the reasoning.
    for name, options in (
        ("MD_CALL", '"P,LP,VUS,LB,B"'),
        ("WHO_IS_RIGHT", '"clinvar,fastvep,neither"'),
    ):
        letter = get_column_letter(header.index(name) + 1)
        dv = DataValidation(type="list", formula1=options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"

    key = wb.create_sheet("key")
    key.column_dimensions["A"].width = 34
    key.column_dimensions["B"].width = 96
    rows_out = [
        ("Every call where fastVEP still commits to a direction ClinVar contradicts", ""),
        ("", ""),
        ("Columns to fill in", ""),
        *REVIEW_COLUMNS,
        ("", ""),
        ("Carried from earlier rounds", ""),
        ("rounds_seen", "which review tables this variant has appeared in"),
        ("md_round2_verdict", "your ruling on the 122-row table, decoded from the cell colours"),
        ("md_round2_note", "your free text on the 122-row table"),
        ("md_round3_note", "the 46-row table; carried your round-2 text forward"),
        ("md_round4_note", "the 122-row table as most recently returned"),
    ]
    for a, b in rows_out:
        key.append([a, b])
    key["A1"].font = Font(bold=True, size=13)
    key["A3"].font = Font(bold=True)
    key[f"A{3 + len(REVIEW_COLUMNS) + 2}"].font = Font(bold=True)
    for row in key.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main() -> None:
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    review_tsv, out_xlsx = (Path(a) for a in sys.argv[1:])

    with review_tsv.open() as f:
        review_rows = list(csv.DictReader(f, delimiter="\t"))

    round2 = load_round2()
    round3 = load_notes(ROUND3_XLSX, "Sheet1", "prior_reviewer_note")
    round4 = load_notes(ROUND4_XLSX, "Sheet1", "prior_new_reviewer_note")

    evidence_columns = [c for c in review_rows[0] if c not in LEAD_COLUMNS]
    header = (
        LEAD_COLUMNS
        + CARRIED_COLUMNS
        + [name for name, _ in REVIEW_COLUMNS]
        + evidence_columns
    )

    rows = []
    for r in review_rows:
        key = key_of(r)
        verdict, note2 = round2.get(key, ("", ""))
        seen = [str(n) for n, src in ((2, round2), (3, round3), (4, round4)) if key in src]
        rows.append({
            **r,
            "rounds_seen": ", ".join(seen),
            "md_round2_verdict": verdict,
            "md_round2_note": note2,
            "md_round3_note": round3.get(key, ""),
            "md_round4_note": round4.get(key, ""),
        })

    # Rows she has never seen lead, then rows she saw but did not rule on, then
    # the ones already ruled on. Highest priority first within each group.
    rows.sort(key=lambda d: (
        bool(d["rounds_seen"]), bool(d["md_round2_verdict"]), -float(d["priority_score"])))

    write_workbook(out_xlsx, rows, header)

    seen = sum(1 for r in rows if r["rounds_seen"])
    ruled = sum(1 for r in rows if r["md_round2_verdict"])
    print(f"{out_xlsx.name}: {len(rows)} still-discordant rows; "
          f"{len(rows) - seen} never reviewed, {seen - ruled} seen but unruled, {ruled} ruled on")


if __name__ == "__main__":
    main()
