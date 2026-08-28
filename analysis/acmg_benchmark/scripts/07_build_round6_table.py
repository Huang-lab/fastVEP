#!/usr/bin/env python3
"""Build the round-6 medical-genetics review workbook.

Same shape as round 5: one sheet of variants, every call where fastVEP still
commits to a direction ClinVar 2-star+ contradicts. Rows she has resolved are
gone by construction, because they are no longer discordant.

Round 6 differs from round 5 in one respect. Round 5's workbook went out and
came back unmarked, so the rows in it are "seen, not ruled on" rather than
"never seen", and `seen_in_round` records which round each row first appeared
in. Her verdicts and notes still come from the rounds she did write in, carried
onto the rows they belong to.

Usage:
  07_build_round6_table.py <review.tsv> <out.xlsx>
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[3]

# Every file the reviewer has written in. The first carries her verdicts as a
# text column; the rest carry her notes forward through our own re-runs.
REVIEWED_TSV = ROOT / "data/benchmark/output_v8/md_review_round2_annotated.tsv"
REVIEWED_XLSX = [
    (ROOT / "data/benchmark/output_v9/discordant_46_round3.xlsx", "prior_reviewer_note"),
    (ROOT / "data/benchmark/discordance_review/discordant_122_round4.xlsx",
     "prior_new_reviewer_note"),
]

# Every workbook that has been sent, oldest first. A row's `seen_in_round` is
# the earliest of these it appeared in; rows in none of them are new this round.
# Round 5 is here because it was sent, not because it was written in.
SENT_XLSX = [
    (ROOT / "data/benchmark/output_v9/discordant_46_round3.xlsx", "round 3"),
    (ROOT / "data/benchmark/discordance_review/discordant_122_round4.xlsx", "round 4"),
    (ROOT / "data/benchmark/discordance_review/discordant_59_round5.xlsx", "round 5"),
]

# Her ruling is recorded as a cell fill in the workbooks she returns and as a
# text column in the TSV; the two agree, and the TSV needs no decoding.
VERDICT_LABEL = {
    "clinvar_correct": "ClinVar correct; fastVEP wrong",
    "fastvep_correct": "fastVEP correct; ClinVar wrong",
    "both_wrong": "both calls discordant with the reviewer's call",
    "needs_literature": "needs literature evidence",
}

# Columns the reviewer fills in this round. The last two are new and exist
# because her round-5 comments are about exactly those two things: whether a
# frequency criterion may be applied to a given disorder at all, and whether
# published functional data should override a prediction.
REVIEW_COLUMNS = [
    ("MD_CALL", "P / LP / VUS / LB / B - your classification for this variant"),
    ("WHO_IS_RIGHT", "clinvar / fastvep / neither"),
    ("REASON_FOR_DISCORDANCE", "free text; what evidence decides it"),
    ("ACMG_CODE_MISAPPLIED", "e.g. BS1, PVS1, BP7 - which code fastVEP got wrong"),
    (
        "BS1_BS2_NOT_APPLICABLE_WHY",
        "if frequency evidence should not apply here: common phenotype / late onset / "
        "reduced penetrance / variable expressivity / founder allele / hypomorph in trans",
    ),
    ("FUNCTIONAL_OR_LITERATURE_PMID", "PMID for functional or segregation data that settles it"),
]

CARRIED_COLUMNS = ["seen_in_round", "md_verdict", "md_note"]

# Identifying columns, pulled to the front so a row can be located at a glance.
LEAD_COLUMNS = [
    "gene", "variant", "fastvep_HGVSc", "fastvep_HGVSp", "stars",
    "truth_class", "fastvep_class", "fastvep_met_criteria",
]

HEADER_FILL = PatternFill("solid", start_color="FF1F3864")
REVIEW_FILL = PatternFill("solid", start_color="FFFFF2CC")
CARRIED_FILL = PatternFill("solid", start_color="FFEAF1F8")


def key_of(values: dict) -> tuple[str, ...]:
    return tuple(str(values[c]).strip() for c in ("chrom", "pos", "ref", "alt"))


def load_annotations() -> dict[tuple[str, ...], dict[str, str]]:
    """Verdict and note per variant, merged across every file she has written in.

    The longest note wins where copies differ, which in practice means the one
    that did not lose a trailing citation to a round trip through Excel.
    """
    out: dict[tuple[str, ...], dict[str, str]] = {}

    def record(key, verdict="", note=""):
        row = out.setdefault(key, {"md_verdict": "", "md_note": ""})
        if verdict:
            row["md_verdict"] = verdict
        if len(note) > len(row["md_note"]):
            row["md_note"] = note

    if REVIEWED_TSV.exists():
        with REVIEWED_TSV.open() as f:
            for r in csv.DictReader(f, delimiter="\t"):
                record(
                    key_of(r),
                    VERDICT_LABEL.get((r.get("reviewer_verdict") or "").strip(), ""),
                    (r.get("REASON FOR DISCORDANCE") or "").strip(),
                )

    for path, column in REVIEWED_XLSX:
        if not path.exists():
            continue
        ws = load_workbook(path)["Sheet1"]
        header = [c.value for c in ws[1]]
        idx = {name: i for i, name in enumerate(header) if name}
        if column not in idx:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["chrom"]] is None:
                continue
            values = {c: row[idx[c]] for c in ("chrom", "pos", "ref", "alt")}
            record(key_of(values), note=(row[idx[column]] or "").strip())

    return out


def load_sent_rounds() -> dict[tuple[str, ...], str]:
    """The earliest sent workbook each variant appeared in, by key."""
    out: dict[tuple[str, ...], str] = {}
    for path, label in SENT_XLSX:
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [h for h in next(rows)]
        idx = {name: i for i, name in enumerate(header) if name}
        if not {"chrom", "pos", "ref", "alt"} <= set(idx):
            continue
        for row in rows:
            if row[idx["chrom"]] is None:
                continue
            key = tuple(str(row[idx[c]]).strip() for c in ("chrom", "pos", "ref", "alt"))
            out.setdefault(key, label)
    return out


def write_workbook(path: Path, rows: list[dict[str, str]], header: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(header)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    review_names = {name for name, _ in REVIEW_COLUMNS}
    for row in rows:
        ws.append([row.get(name, "") for name in header])

    for i, name in enumerate(header, start=1):
        letter = get_column_letter(i)
        if name in review_names:
            width = 26
        elif name.startswith("md_"):
            width = 34
        elif name in ("gene", "stars", "seen_in_round"):
            width = 13
        elif name in ("truth_class", "fastvep_class"):
            width = 17
        else:
            width = 22
        ws.column_dimensions[letter].width = width
        fill = REVIEW_FILL if name in review_names else (
            CARRIED_FILL if name in CARRIED_COLUMNS else None)
        if fill:
            for cell in ws[letter][1:]:
                cell.fill = fill

    # Dropdowns on the two columns with a closed vocabulary. Everything else is
    # free text on purpose: the useful part of her review is the reasoning.
    for name, options in (
        ("MD_CALL", '"P,LP,VUS,LB,B"'),
        ("WHO_IS_RIGHT", '"clinvar,fastvep,neither"'),
    ):
        letter = get_column_letter(header.index(name) + 1)
        dv = DataValidation(type="list", formula1=options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"

    key = wb.create_sheet("key")
    key.column_dimensions["A"].width = 34
    key.column_dimensions["B"].width = 96
    rows_out = [
        ("Every call where fastVEP still commits to a direction ClinVar contradicts", ""),
        ("", ""),
        ("Columns to fill in", ""),
        *REVIEW_COLUMNS,
        ("", ""),
        ("Carried from your review", ""),
        ("seen_in_round", "the earliest round whose table carried this row; blank means new"),
        ("md_verdict", "your ruling, decoded from the cell colours you used"),
        ("md_note", "your free text, verbatim"),
    ]
    for a, b in rows_out:
        key.append([a, b])
    key["A1"].font = Font(bold=True, size=13)
    key["A3"].font = Font(bold=True)
    key[f"A{3 + len(REVIEW_COLUMNS) + 2}"].font = Font(bold=True)
    for row in key.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main() -> None:
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    review_tsv, out_xlsx = (Path(a) for a in sys.argv[1:])

    with review_tsv.open() as f:
        review_rows = list(csv.DictReader(f, delimiter="\t"))

    annotations = load_annotations()

    evidence_columns = [c for c in review_rows[0] if c not in LEAD_COLUMNS]
    header = (
        LEAD_COLUMNS
        + CARRIED_COLUMNS
        + [name for name, _ in REVIEW_COLUMNS]
        + evidence_columns
    )

    seen = load_sent_rounds()
    rows = []
    for r in review_rows:
        k = key_of(r)
        prior = annotations.get(k)
        rows.append({
            **r,
            "seen_in_round": seen.get(k, ""),
            "md_verdict": prior["md_verdict"] if prior else "",
            "md_note": prior["md_note"] if prior else "",
        })

    # Rows she has never seen lead, then rows she saw but did not rule on, then
    # the ones already ruled on. Highest priority first within each group.
    rows.sort(key=lambda d: (
        bool(d["seen_in_round"]), bool(d["md_verdict"]), -float(d["priority_score"])))

    write_workbook(out_xlsx, rows, header)

    seen = sum(1 for r in rows if r["seen_in_round"])
    ruled = sum(1 for r in rows if r["md_verdict"])
    print(f"{out_xlsx.name}: {len(rows)} still-discordant rows; "
          f"{len(rows) - seen} never sent before, {seen - ruled} sent but unruled, {ruled} ruled on")


if __name__ == "__main__":
    main()
