#!/usr/bin/env python3
"""Build the round-5 medical-genetics review workbooks.

Two files, from one enriched review table:

  strong_discordant_round5.xlsx   every current opposite-direction call, with
                                  the reviewer's round-4 verdict and note
                                  carried across where she has already ruled
  new_for_review_round5.xlsx      only the rows she has never seen, with the
                                  same schema and the judgement columns blank

The split is the point. The first file is the standing record of where fastVEP
and ClinVar disagree in direction, so a row she already adjudicated stays
visible with her ruling attached rather than being asked again. The second is
the work: rows that have appeared since her last pass, or that her pass never
covered.

Her round-4 ruling is recorded as cell fill colour rather than text, keyed by
the legend on that workbook's second sheet. `VERDICTS` below decodes it. Both
the decoded verdict and her free-text note travel into the new tables as
ordinary columns, so nothing depends on colour surviving another round trip.

Usage:
  06_build_round5_tables.py <review.tsv> <round4.xlsx> <out-dir>
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Round-4 legend, from Sheet2 of the returned workbook. The key is
# (column the fill sits on, fill colour); theme colours are keyed by index
# because openpyxl reports them without an RGB value.
VERDICTS = {
    ("truth_class", "FF92D050"): "ClinVar correct; fastVEP wrong",
    ("fastvep_class", "FF92D050"): "fastVEP correct; ClinVar wrong",
    ("truth_class", "theme8"): "both calls discordant with the reviewer's call",
    ("fastvep_class", "theme8"): "both calls discordant with the reviewer's call",
    ("truth_class", "theme5"): "needs literature evidence",
    ("fastvep_class", "theme5"): "needs literature evidence",
    ("truth_class", "FFFFC000"): "homology gene; population frequencies unreliable (PMID 27228465)",
    ("fastvep_class", "FFFFC000"): "homology gene; population frequencies unreliable (PMID 27228465)",
    ("truth_class", "FFFFFF00"): "does '.' mean del?",
    ("fastvep_class", "FFFFFF00"): "does '.' mean del?",
}

# Columns the reviewer fills in. The last two are new this round and exist
# because her round-5 comments are about exactly those two things: whether a
# frequency criterion may be applied to a given disorder at all, and whether
# published functional data should override a prediction.
REVIEW_COLUMNS = [
    ("MD_CALL", "P / LP / VUS / LB / B - your classification for this variant"),
    ("WHO_IS_RIGHT", "clinvar / fastvep / neither"),
    ("REASON_FOR_DISCORDANCE", "free text; what evidence decides it"),
    ("ACMG_CODE_MISAPPLIED", "e.g. BS1, PVS1, BP7 - which code fastVEP got wrong"),
    (
        "BS1_BS2_NOT_APPLICABLE_WHY",
        "if frequency evidence should not apply here: common phenotype / late onset / "
        "reduced penetrance / variable expressivity / founder allele / hypomorph in trans",
    ),
    ("FUNCTIONAL_OR_LITERATURE_PMID", "PMID for functional or segregation data that settles it"),
]

# Carried across from round 4, shown before the evidence.
CARRIED_COLUMNS = ["reviewed_in_round4", "md_round4_verdict", "md_round4_note"]

# Identifying columns, pulled to the front so a row can be located at a glance.
LEAD_COLUMNS = [
    "gene", "variant", "fastvep_HGVSc", "fastvep_HGVSp", "stars",
    "truth_class", "fastvep_class", "fastvep_met_criteria",
]

HEADER_FILL = PatternFill("solid", start_color="FF1F3864")
REVIEW_FILL = PatternFill("solid", start_color="FFFFF2CC")
CARRIED_FILL = PatternFill("solid", start_color="FFEAF1F8")


def fill_key(cell) -> str | None:
    """A stable string for a cell's fill, or None when it has no fill."""
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    colour = fill.start_color
    if colour.type == "rgb" and isinstance(colour.rgb, str):
        return colour.rgb
    if colour.type == "theme":
        try:
            return f"theme{int(colour.theme)}"
        except (TypeError, ValueError):
            return None
    return None


def load_round4(path: Path) -> dict[tuple[str, ...], dict[str, str]]:
    """Decode the reviewer's round-4 rulings, keyed by (chrom, pos, ref, alt)."""
    ws = load_workbook(path)["Sheet1"]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header) if name}
    out: dict[tuple[str, ...], dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        if values[idx["chrom"]] is None:
            continue
        verdicts: list[str] = []
        for column in ("truth_class", "fastvep_class"):
            decoded = VERDICTS.get((column, fill_key(row[idx[column]])))
            if decoded and decoded not in verdicts:
                verdicts.append(decoded)
        key = tuple(str(values[idx[c]]).strip() for c in ("chrom", "pos", "ref", "alt"))
        note = values[idx.get("prior_new_reviewer_note", -1)] if "prior_new_reviewer_note" in idx else None
        out[key] = {
            "md_round4_verdict": "; ".join(verdicts),
            "md_round4_note": (note or "").strip(),
        }
    return out


def write_workbook(path: Path, rows: list[dict[str, str]], header: list[str], title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(header)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    review_names = {name for name, _ in REVIEW_COLUMNS}
    for row in rows:
        ws.append([row.get(name, "") for name in header])

    for i, name in enumerate(header, start=1):
        letter = get_column_letter(i)
        if name in review_names:
            width = 26
        elif name in CARRIED_COLUMNS:
            width = 30
        elif name in ("gene", "stars"):
            width = 10
        elif name in ("truth_class", "fastvep_class"):
            width = 17
        else:
            width = 22
        ws.column_dimensions[letter].width = width
        fill = REVIEW_FILL if name in review_names else (
            CARRIED_FILL if name in CARRIED_COLUMNS else None)
        if fill:
            for cell in ws[letter][1:]:
                cell.fill = fill

    # Dropdowns on the two columns with a closed vocabulary. Everything else is
    # free text on purpose: the useful part of her review is the reasoning.
    for name, options in (
        ("MD_CALL", '"P,LP,VUS,LB,B"'),
        ("WHO_IS_RIGHT", '"clinvar,fastvep,neither"'),
    ):
        if name not in header:
            continue
        letter = get_column_letter(header.index(name) + 1)
        dv = DataValidation(type="list", formula1=options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"

    key = wb.create_sheet("key")
    key.column_dimensions["A"].width = 34
    key.column_dimensions["B"].width = 96
    key.append([title, ""])
    key["A1"].font = Font(bold=True, size=13)
    key.append(["", ""])
    key.append(["Column", "What to put in it"])
    for cell in key[3]:
        cell.font = Font(bold=True)
    for name, description in REVIEW_COLUMNS:
        key.append([name, description])
    key.append(["", ""])
    key.append(["Carried from round 4", ""])
    key[len(key["A"])][0].font = Font(bold=True)
    key.append(["md_round4_verdict", "your ruling last round, decoded from the cell colours"])
    key.append(["md_round4_note", "your free-text note last round, verbatim"])
    for row in key.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main() -> None:
    if len(sys.argv) != 4:
        sys.exit(__doc__)
    review_tsv, round4_xlsx, out_dir = (Path(a) for a in sys.argv[1:])
    out_dir.mkdir(parents=True, exist_ok=True)

    with review_tsv.open() as f:
        review_rows = list(csv.DictReader(f, delimiter="\t"))
    round4 = load_round4(round4_xlsx)

    evidence_columns = [c for c in review_rows[0] if c not in LEAD_COLUMNS]
    header = (
        LEAD_COLUMNS
        + CARRIED_COLUMNS
        + [name for name, _ in REVIEW_COLUMNS]
        + evidence_columns
    )

    rows = []
    for r in review_rows:
        key = (r["chrom"], r["pos"], r["ref"], r["alt"])
        prior = round4.get(key)
        row = dict(r)
        row["reviewed_in_round4"] = "yes" if prior else "no"
        row["md_round4_verdict"] = prior["md_round4_verdict"] if prior else ""
        row["md_round4_note"] = prior["md_round4_note"] if prior else ""
        rows.append(row)

    # Sort so the unruled rows lead, highest priority first within each group:
    # the reviewer's time goes to what has never been adjudicated.
    rows.sort(key=lambda d: (bool(d["md_round4_verdict"]), -float(d["priority_score"])))

    write_workbook(
        out_dir / "strong_discordant_round5.xlsx", rows, header,
        "All current opposite-direction calls, with round-4 rulings carried across")

    # "Remaining" means unruled, not unseen: a row that appeared in the round-4
    # workbook but carries no colour was never adjudicated, and asking about it
    # again is the point of sending a second file.
    unruled = [r for r in rows if not r["md_round4_verdict"]]
    write_workbook(
        out_dir / "new_for_review_round5.xlsx", unruled, header,
        "Opposite-direction calls with no ruling from the round-4 review")

    seen = sum(1 for r in rows if r["reviewed_in_round4"] == "yes")
    print(f"strong_discordant_round5.xlsx  {len(rows)} rows "
          f"({seen} appeared in the round-4 workbook, {len(rows) - len(unruled)} carry a ruling)")
    print(f"new_for_review_round5.xlsx     {len(unruled)} rows")


if __name__ == "__main__":
    main()
